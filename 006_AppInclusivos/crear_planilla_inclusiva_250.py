from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import datetime

wb = Workbook()

AZUL_OSC   = "1F3864"
AZUL_MED   = "2E75B6"
AZUL_FONDO = "DEEAF1"
VERDE      = "375623"
VERDE_CLAR = "E2EFDA"
ROJO       = "C00000"
ROJO_CLAR  = "FCE4D6"
NARANJA    = "C55A11"
NARANJA_C  = "FCE9D9"
AMARILLO   = "FFF2CC"
GRIS_ALT   = "F2F2F2"
BLANCO     = "FFFFFF"

MAX_ALUMNOS  = 250   # filas en hoja Alumnos
MAX_FAM      = 750   # filas en hoja Familias (3 compromisos/alumno aprox.)
MAX_SEG      = 750   # filas en hoja Seguimiento
MAX_MEC      = 80    # filas en hoja Documentos MEC

def fill(h): return PatternFill("solid", fgColor=h)
def hdr_font(color=BLANCO): return Font(name="Arial", bold=True, color=color, size=10)
def cell_font(bold=False, color="000000"): return Font(name="Arial", bold=bold, color=color, size=10)
def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def write_header_row(ws, row, headers_widths):
    for i, (h, _) in enumerate(headers_widths, 1):
        c = ws.cell(row=row, column=i)
        c.value = h; c.font = hdr_font(); c.fill = fill(AZUL_MED)
        c.alignment = center(); c.border = border_thin()
    ws.row_dimensions[row].height = 30

def fill_data_rows(ws, start, end, ncols, formulas=None):
    for r in range(start, end + 1):
        bg = GRIS_ALT if r % 2 == 0 else BLANCO
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.fill = fill(bg); cell.font = cell_font()
            cell.alignment = left(); cell.border = border_thin()
        ws.cell(r, 1).value = r - start + 1
        ws.cell(r, 1).alignment = center()
        ws.row_dimensions[r].height = 18
        if formulas:
            for col, fmt in formulas.items():
                ws.cell(r, col).value = fmt.format(r=r)
                ws.cell(r, col).alignment = center()

def add_cond_fmt(ws, rng):
    rojo_f  = PatternFill("solid", fgColor=ROJO_CLAR)
    nar_f   = PatternFill("solid", fgColor=NARANJA_C)
    amar_f  = PatternFill("solid", fgColor=AMARILLO)
    verde_f = PatternFill("solid", fgColor=VERDE_CLAR)
    ws.conditional_formatting.add(rng, CellIsRule("lessThan",    ["0"],     fill=rojo_f,  font=Font(name="Arial",color=ROJO,  bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule("between",     ["0","15"],fill=nar_f,   font=Font(name="Arial",color=NARANJA,bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule("between",     ["16","30"],fill=amar_f))
    ws.conditional_formatting.add(rng, CellIsRule("greaterThan", ["30"],    fill=verde_f, font=Font(name="Arial",color=VERDE)))

def title_row(ws, text, ncols, bg=AZUL_OSC):
    ws.merge_cells(f"A1:{chr(64+ncols)}1")
    c = ws["A1"]
    c.value = text; c.font = Font(name="Arial",bold=True,color=BLANCO,size=13)
    c.fill = fill(bg); c.alignment = center()
    ws.row_dimensions[1].height = 30

# ══════════════════════════════════════════════════════════
# HOJA 1: INICIO
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 INICIO"
ws1.sheet_view.showGridLines = False
for col,w in zip("ABCDEFGH",[3,28,22,16,16,20,18,3]): ws1.column_dimensions[col].width=w

ws1.merge_cells("B2:G3")
t=ws1["B2"]; t.value="COLEGIO PRESBITERIANO CERRITOS"
t.font=Font(name="Arial",bold=True,color=BLANCO,size=16); t.fill=fill(AZUL_OSC); t.alignment=center()
ws1.row_dimensions[1].height=8; ws1.row_dimensions[2].height=22; ws1.row_dimensions[3].height=22

ws1.merge_cells("B4:G4")
s=ws1["B4"]; s.value="Sistema de Seguimiento — Educación Inclusiva"
s.font=Font(name="Arial",italic=True,color=AZUL_MED,size=12); s.alignment=center()
ws1.row_dimensions[5].height=8

ws1.merge_cells("B6:C6")
ws1["B6"].value="RESUMEN GENERAL"; ws1["B6"].font=Font(name="Arial",bold=True,color=BLANCO,size=10)
ws1["B6"].fill=fill(AZUL_MED); ws1["B6"].alignment=center()

stats = [
    ("Total alumnos registrados",      f"=COUNTA('📚 Alumnos'!B4:B{MAX_ALUMNOS+3})"),
    ("Compromisos activos (Pendiente)", f"=COUNTIF('🤝 Familias'!F4:F{MAX_FAM+3},\"Pendiente\")"),
    ("Documentos MEC pendientes",       f"=COUNTIF('📄 Documentos MEC'!E4:E{MAX_MEC+3},\"Pendiente\")"),
    ("Alertas próximas (≤30 días)",     '=COUNTIF(\'📋 INICIO\'!F15:F114,"⚠️ PRÓXIMO")+COUNTIF(\'📋 INICIO\'!F15:F114,"🔴 VENCIDO")'),
]
for i,(label,formula) in enumerate(stats):
    r=7+i; bg=AZUL_FONDO
    ws1.cell(r,2).value=label;   ws1.cell(r,2).fill=fill(bg); ws1.cell(r,2).font=cell_font(); ws1.cell(r,2).border=border_thin(); ws1.cell(r,2).alignment=left()
    ws1.cell(r,3).value=formula; ws1.cell(r,3).fill=fill(bg); ws1.cell(r,3).font=Font(name="Arial",bold=True,color=AZUL_MED,size=12); ws1.cell(r,3).border=border_thin(); ws1.cell(r,3).alignment=center()

ws1.row_dimensions[11].height=8
ws1.merge_cells("B12:G12")
ph=ws1["B12"]; ph.value="⏰  PRÓXIMOS VENCIMIENTOS Y ALERTAS"
ph.font=Font(name="Arial",bold=True,color=BLANCO,size=11); ph.fill=fill(AZUL_OSC); ph.alignment=center()

for i,h in enumerate(["Tipo","Alumno / Descripción","Fecha Límite","Días Restantes","Estado","Responsable"],2):
    c=ws1.cell(13,i); c.value=h; c.font=hdr_font(); c.fill=fill(AZUL_MED); c.alignment=center(); c.border=border_thin()
ws1.row_dimensions[13].height=28

sample=[
    ("📄 Doc. MEC","Planilla de matrícula inclusiva","2026-07-31"),
    ("📄 Doc. MEC","Informe semestral de adecuaciones","2026-07-15"),
    ("📄 Doc. MEC","Plan Educativo Individual (PEI)","2026-06-30"),
    ("🤝 Familia","Reunión de compromiso — ver lista",""),
    ("📄 Doc. MEC","Evaluación psicopedagógica","2026-08-30"),
]
for i,(tipo,desc,fecha) in enumerate(sample):
    r=14+i; bg=GRIS_ALT if i%2==0 else BLANCO
    for col in [2,3,4,5,6,7]:
        ws1.cell(r,col).fill=fill(bg); ws1.cell(r,col).border=border_thin(); ws1.cell(r,col).font=cell_font(); ws1.cell(r,col).alignment=center()
    ws1.cell(r,2).value=tipo; ws1.cell(r,3).value=desc; ws1.cell(r,3).alignment=left()
    if fecha:
        ws1.cell(r,4).value=fecha; ws1.cell(r,4).number_format="DD/MM/YYYY"
        ws1.cell(r,5).value=f"=IF(D{r}=\"\",\"\",D{r}-TODAY())"; ws1.cell(r,5).number_format="0"
        ws1.cell(r,6).value=f'=IF(D{r}="","",IF(E{r}<0,"🔴 VENCIDO",IF(E{r}<=15,"⚠️ PRÓXIMO",IF(E{r}<=30,"🟡 ATENCIÓN","🟢 OK"))))'
    ws1.cell(r,7).value="Lic. Rodrigo Godoy"

ws1.merge_cells("B20:G20")
ws1["B20"].value="ℹ  Completá los datos en las hojas 📚 Alumnos, 🤝 Familias y 📄 Documentos MEC — las alertas se actualizan automáticamente."
ws1["B20"].font=Font(name="Arial",italic=True,color="595959",size=9); ws1["B20"].alignment=left()
ws1.freeze_panes="B14"

# ══════════════════════════════════════════════════════════
# HOJA 2: ALUMNOS (250 filas)
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📚 Alumnos")
ws2.sheet_view.showGridLines=False
HW_ALU=[("N°",5),("Apellido y Nombre",28),("CI / Doc.",14),("Fecha Nacim.",14),
         ("Curso / Grado",14),("Turno",10),("Necesidad / Diagnóstico",30),
         ("Tipo de Adecuación",22),("Grado de Apoyo",16),("Fecha Ingreso",14),
         ("Docente Responsable",24),("Tel. Familia",16),("Responsable MEC",22),
         ("Estado",14),("Observaciones",35)]
for i,(h,w) in enumerate(HW_ALU,1): ws2.column_dimensions[chr(64+i)].width=w

title_row(ws2,"REGISTRO DE ALUMNOS INCLUSIVOS — Colegio Presbiteriano Cerritos (hasta 250 alumnos)",len(HW_ALU))
ws2.merge_cells(f"A2:{chr(64+len(HW_ALU))}2")
ws2["A2"].value=f'Actualizado: {datetime.date.today().strftime("%d/%m/%Y")}  |  Normativa: Ley 5136/2013 | Res. MEC 29.664/2012 | Decreto 1.350/2019'
ws2["A2"].font=Font(name="Arial",italic=True,color=AZUL_MED,size=9); ws2["A2"].alignment=left()
write_header_row(ws2,3,HW_ALU)

dv_t=DataValidation(type="list",formula1='"Mañana,Tarde,Tiempo Completo"',allow_blank=True)
dv_a=DataValidation(type="list",formula1='"Adecuación de Acceso,Adecuación Curricular No Significativa,Adecuación Curricular Significativa"',allow_blank=True)
dv_g=DataValidation(type="list",formula1='"Leve,Moderado,Intenso"',allow_blank=True)
dv_e=DataValidation(type="list",formula1='"Activo,Egresado,Trasladado,Retirado"',allow_blank=True)
ws2.add_data_validation(dv_t); ws2.add_data_validation(dv_a)
ws2.add_data_validation(dv_g); ws2.add_data_validation(dv_e)

END_ALU = MAX_ALUMNOS + 3
for r in range(4, END_ALU + 1):
    bg=GRIS_ALT if r%2==0 else BLANCO
    for c in range(1,len(HW_ALU)+1):
        cell=ws2.cell(r,c); cell.fill=fill(bg); cell.font=cell_font(); cell.alignment=left(); cell.border=border_thin()
    ws2.cell(r,1).value=r-3; ws2.cell(r,1).alignment=center()
    ws2.cell(r,4).number_format="DD/MM/YYYY"
    ws2.cell(r,10).number_format="DD/MM/YYYY"
    dv_t.add(f"F{r}"); dv_a.add(f"H{r}"); dv_g.add(f"I{r}"); dv_e.add(f"N{r}")
    ws2.row_dimensions[r].height=18

ws2.freeze_panes="A4"

# ══════════════════════════════════════════════════════════
# HOJA 3: FAMILIAS (750 filas)
# ══════════════════════════════════════════════════════════
ws3=wb.create_sheet("🤝 Familias")
ws3.sheet_view.showGridLines=False
HW_FAM=[("N°",5),("Alumno",28),("Fecha Reunión",14),("Participantes",30),
         ("Acuerdo / Compromiso",40),("Estado",14),("Fecha Límite",14),
         ("Días Restantes",14),("Seguimiento",35),("Próxima Reunión",16),
         ("Responsable",22),("Firma / Constancia",18)]
for i,(h,w) in enumerate(HW_FAM,1): ws3.column_dimensions[chr(64+i)].width=w

title_row(ws3,"COMPROMISOS Y ACUERDOS CON FAMILIAS — Educación Inclusiva (hasta 750 registros)",len(HW_FAM))
ws3.merge_cells(f"A2:{chr(64+len(HW_FAM))}2")
ws3["A2"].value="Registrá cada reunión y acuerdo formal con la familia. Los Días Restantes se calculan automáticamente."
ws3["A2"].font=Font(name="Arial",italic=True,color=AZUL_MED,size=9)
write_header_row(ws3,3,HW_FAM)

dv_ef=DataValidation(type="list",formula1='"Pendiente,Cumplido,Reprogramado,Incumplido"',allow_blank=True)
ws3.add_data_validation(dv_ef)

END_FAM = MAX_FAM + 3
for r in range(4, END_FAM + 1):
    bg=GRIS_ALT if r%2==0 else BLANCO
    for c in range(1,len(HW_FAM)+1):
        cell=ws3.cell(r,c); cell.fill=fill(bg); cell.font=cell_font(); cell.alignment=left(); cell.border=border_thin()
    ws3.cell(r,1).value=r-3; ws3.cell(r,1).alignment=center()
    ws3.cell(r,3).number_format="DD/MM/YYYY"
    ws3.cell(r,7).number_format="DD/MM/YYYY"
    ws3.cell(r,8).value=f'=IF(G{r}="","",G{r}-TODAY())'; ws3.cell(r,8).number_format="0"
    ws3.cell(r,8).alignment=center()
    ws3.cell(r,10).number_format="DD/MM/YYYY"
    dv_ef.add(f"F{r}")
    ws3.row_dimensions[r].height=18

add_cond_fmt(ws3, f"H4:H{END_FAM}")
ws3.freeze_panes="A4"

# ══════════════════════════════════════════════════════════
# HOJA 4: DOCUMENTOS MEC (80 filas)
# ══════════════════════════════════════════════════════════
ws4=wb.create_sheet("📄 Documentos MEC")
ws4.sheet_view.showGridLines=False
HW_MEC=[("N°",5),("Documento / Trámite",40),("Alumno(s)",28),("Base Legal",30),
         ("Estado",14),("Fecha Límite",14),("Días Restantes",14),("Alerta",14),
         ("Responsable",22),("Dónde se presenta",25),("Observaciones",35)]
for i,(h,w) in enumerate(HW_MEC,1): ws4.column_dimensions[chr(64+i)].width=w

title_row(ws4,"DOCUMENTOS Y PLAZOS — MINISTERIO DE EDUCACIÓN Y CIENCIAS (MEC) — Paraguay",len(HW_MEC))
ws4.merge_cells(f"A2:{chr(64+len(HW_MEC))}2")
ws4["A2"].value="Normativa: Res. MEC N° 29.664/2012 | Decreto N° 1.350/2019 | Ley N° 5136/2013 | Res. N° 29.451/2010"
ws4["A2"].font=Font(name="Arial",italic=True,color=AZUL_MED,size=9)
write_header_row(ws4,3,HW_MEC)

dv_em=DataValidation(type="list",formula1='"Pendiente,En preparación,Presentado,Aprobado,Observado"',allow_blank=True)
ws4.add_data_validation(dv_em)

docs=[
    ("Planilla de matrícula inclusiva (Ficha de Registro)","Todos los alumnos nuevos","Res. MEC 29.664/2012 Art. 8","Pendiente","2026-03-15","Sección Ed. Especial MEC / Supervisión"),
    ("Plan Educativo Individual (PEI) — Elaboración","Cada alumno inclusivo","Ley 5136/2013 Art. 15","En preparación","2026-04-30","Archivo institucional + copia Supervisión"),
    ("Informe de adecuaciones curriculares — 1er semestre","Todos los alumnos inclusivos","Res. MEC 29.664/2012 Art. 12","Pendiente","2026-07-15","Supervisión Educativa Departamental"),
    ("Informe de adecuaciones curriculares — 2do semestre","Todos los alumnos inclusivos","Res. MEC 29.664/2012 Art. 12","Pendiente","2026-11-30","Supervisión Educativa Departamental"),
    ("Evaluación psicopedagógica actualizada","Alumnos que requieren renovación","Decreto 1.350/2019 Art. 22","Pendiente","2026-08-31","Centro de Diagnóstico MEC / archivo"),
    ("Nómina de alumnos inclusivos por nivel","Todos los niveles","Res. MEC 29.664/2012","Pendiente","2026-03-31","Supervisión Educativa"),
    ("Constancia de diagnóstico médico/psicológico","Alumnos nuevos","Ley 5136/2013 Art. 9","Pendiente","","Archivo institucional"),
    ("Plan de Apoyos y Recursos (PAR)","Adecuación significativa","Decreto 1.350/2019","Pendiente","2026-05-15","Archivo institucional"),
    ("Acta de compromiso familiar firmada","Cada familia de alumno inclusivo","Res. MEC 29.664/2012","Pendiente","","Archivo institucional"),
    ("Informe final anual de progreso","Todos los alumnos inclusivos","Ley 5136/2013 Art. 16","Pendiente","2026-11-30","Supervisión Educativa Departamental"),
    ("Solicitud de recursos de apoyo (intérprete, asistente)","Según necesidad","Decreto 1.350/2019 Art. 30","Pendiente","2026-03-01","DGEEI - Dir. Gral. Ed. Especial e Inclusiva"),
    ("Certificado de discapacidad SENADIS actualizado","Alumnos que corresponda","Ley 4962/2013","Pendiente","","Archivo institucional"),
]
for ri,(doc,alumnos,legal,estado,fecha,lugar) in enumerate(docs):
    r=ri+4; bg=GRIS_ALT if ri%2==0 else BLANCO
    for c in range(1,len(HW_MEC)+1):
        cell=ws4.cell(r,c); cell.fill=fill(bg); cell.font=cell_font(); cell.alignment=left(); cell.border=border_thin()
    ws4.cell(r,1).value=ri+1; ws4.cell(r,1).alignment=center()
    ws4.cell(r,2).value=doc; ws4.cell(r,3).value=alumnos; ws4.cell(r,4).value=legal
    ws4.cell(r,5).value=estado; dv_em.add(f"E{r}")
    if fecha: ws4.cell(r,6).value=fecha; ws4.cell(r,6).number_format="DD/MM/YYYY"
    ws4.cell(r,7).value=f'=IF(F{r}="","",F{r}-TODAY())'; ws4.cell(r,7).number_format="0"; ws4.cell(r,7).alignment=center()
    ws4.cell(r,8).value=f'=IF(F{r}="","",IF(G{r}<0,"🔴 VENCIDO",IF(G{r}<=15,"⚠️ URGENTE",IF(G{r}<=30,"🟡 PRÓXIMO","🟢 OK"))))'; ws4.cell(r,8).alignment=center()
    ws4.cell(r,9).value="Lic. Rodrigo Godoy"; ws4.cell(r,10).value=lugar
    ws4.row_dimensions[r].height=22

END_MEC = MAX_MEC + 3
for r in range(len(docs)+4, END_MEC+1):
    bg=GRIS_ALT if r%2==0 else BLANCO
    for c in range(1,len(HW_MEC)+1):
        cell=ws4.cell(r,c); cell.fill=fill(bg); cell.font=cell_font(); cell.alignment=left(); cell.border=border_thin()
    ws4.cell(r,1).value=r-3; ws4.cell(r,1).alignment=center()
    ws4.cell(r,7).value=f'=IF(F{r}="","",F{r}-TODAY())'; ws4.cell(r,7).number_format="0"; ws4.cell(r,7).alignment=center()
    ws4.cell(r,8).value=f'=IF(F{r}="","",IF(G{r}<0,"🔴 VENCIDO",IF(G{r}<=15,"⚠️ URGENTE",IF(G{r}<=30,"🟡 PRÓXIMO","🟢 OK"))))'; ws4.cell(r,8).alignment=center()
    dv_em.add(f"E{r}")
    ws4.row_dimensions[r].height=18

add_cond_fmt(ws4, f"G4:G{END_MEC}")
ws4.freeze_panes="A4"

# ══════════════════════════════════════════════════════════
# HOJA 5: SEGUIMIENTO (750 filas)
# ══════════════════════════════════════════════════════════
ws5=wb.create_sheet("📊 Seguimiento")
ws5.sheet_view.showGridLines=False
HW_SEG=[("N°",5),("Alumno",28),("Materia / Área",22),("Trimestre / Período",18),
         ("Tipo de Adecuación",24),("Objetivo Planteado",35),("Estrategia Usada",30),
         ("Resultado / Logro",30),("% Logro",10),("Próximo Paso",30),
         ("Fecha Registro",14),("Docente",22)]
for i,(h,w) in enumerate(HW_SEG,1): ws5.column_dimensions[chr(64+i)].width=w

title_row(ws5,"SEGUIMIENTO ACADÉMICO INDIVIDUAL — Alumnos Inclusivos (hasta 750 registros)",len(HW_SEG))
ws5.merge_cells(f"A2:{chr(64+len(HW_SEG))}2")
ws5["A2"].value="Registrá el progreso por alumno, materia y período. El % de logro activa colores automáticamente."
ws5["A2"].font=Font(name="Arial",italic=True,color=AZUL_MED,size=9)
write_header_row(ws5,3,HW_SEG)

dv_trim=DataValidation(type="list",formula1='"1er Trimestre,2do Trimestre,3er Trimestre,Semestral,Anual"',allow_blank=True)
ws5.add_data_validation(dv_trim)

END_SEG = MAX_SEG + 3
for r in range(4, END_SEG+1):
    bg=GRIS_ALT if r%2==0 else BLANCO
    for c in range(1,len(HW_SEG)+1):
        cell=ws5.cell(r,c); cell.fill=fill(bg); cell.font=cell_font(); cell.alignment=left(); cell.border=border_thin()
    ws5.cell(r,1).value=r-3; ws5.cell(r,1).alignment=center()
    ws5.cell(r,9).number_format="0%"; ws5.cell(r,9).alignment=center()
    ws5.cell(r,11).number_format="DD/MM/YYYY"
    dv_trim.add(f"D{r}")
    ws5.row_dimensions[r].height=20

rojo_f=PatternFill("solid",fgColor=ROJO_CLAR); amar_f=PatternFill("solid",fgColor=AMARILLO); verde_f=PatternFill("solid",fgColor=VERDE_CLAR)
ws5.conditional_formatting.add(f"I4:I{END_SEG}", CellIsRule("lessThan",["0.4"],fill=rojo_f,font=Font(name="Arial",color=ROJO,bold=True)))
ws5.conditional_formatting.add(f"I4:I{END_SEG}", CellIsRule("between",["0.4","0.6"],fill=amar_f))
ws5.conditional_formatting.add(f"I4:I{END_SEG}", CellIsRule("greaterThan",["0.6"],fill=verde_f,font=Font(name="Arial",color=VERDE)))
ws5.freeze_panes="A4"

# ══════════════════════════════════════════════════════════
# HOJA 6: CALENDARIO
# ══════════════════════════════════════════════════════════
ws6=wb.create_sheet("📅 Calendario 2026")
ws6.sheet_view.showGridLines=False
HW_CAL=[("Mes",12),("Acción / Evento",45),("Fecha",14),("Normativa",35),("Destinatario",25),("Estado",14),("Días Restantes",14)]
for i,(h,w) in enumerate(HW_CAL,1): ws6.column_dimensions[chr(64+i)].width=w

title_row(ws6,"CALENDARIO ANUAL DE EDUCACIÓN INCLUSIVA 2026 — MEC Paraguay",len(HW_CAL))
write_header_row(ws6,2,HW_CAL)

cal=[
    ("Febrero","Inicio de año — Identificación de alumnos inclusivos","2026-02-16","Ley 5136/2013","Dirección / Docentes"),
    ("Febrero","Entrega informes alumnos inclusivos año anterior","2026-02-28","Res. 29.664/2012","Dirección"),
    ("Marzo","Matrícula especial / diferenciada — Cierre","2026-03-15","Res. 29.664/2012 Art.8","Secretaría"),
    ("Marzo","Nómina de alumnos inclusivos a Supervisión","2026-03-31","Res. 29.664/2012","Dirección"),
    ("Abril","Elaboración PEI — plazo","2026-04-30","Ley 5136/2013 Art.15","Docentes + Psicología"),
    ("Mayo","Plan de Apoyos y Recursos (PAR) — entrega","2026-05-15","Decreto 1.350/2019","Coordinadores"),
    ("Mayo","Reunión con familias — 1er seguimiento","2026-05-30","Res. 29.664/2012","Dirección / Docentes"),
    ("Junio","Evaluación trimestral con adecuaciones — 1er trim.","2026-06-15","Ley 5136/2013 Art.16","Docentes"),
    ("Julio","Informe adecuaciones 1er semestre — Supervisión","2026-07-15","Res. 29.664/2012 Art.12","Dirección"),
    ("Julio","Planilla matrícula inclusiva — actualización","2026-07-31","Res. 29.664/2012","Secretaría"),
    ("Agosto","Evaluaciones psicopedagógicas — renovación","2026-08-31","Decreto 1.350/2019 Art.22","Psicología"),
    ("Septiembre","Revisión PEI — ajuste 2do semestre","2026-09-15","Ley 5136/2013","Docentes + Psicología"),
    ("Septiembre","Reunión con familias — 2do seguimiento","2026-09-30","Res. 29.664/2012","Dirección / Docentes"),
    ("Octubre","Evaluación trimestral con adecuaciones — 2do trim.","2026-10-15","Ley 5136/2013 Art.16","Docentes"),
    ("Noviembre","Informe adecuaciones 2do semestre — Supervisión","2026-11-30","Res. 29.664/2012 Art.12","Dirección"),
    ("Noviembre","Informe final anual de progreso inclusivo","2026-11-30","Ley 5136/2013 Art.16","Dirección"),
    ("Diciembre","Reunión de cierre con familias","2026-12-05","Res. 29.664/2012","Dirección"),
    ("Diciembre","Archivo de expedientes inclusivos del año","2026-12-15","Res. 29.664/2012","Secretaría"),
]
dv_ec=DataValidation(type="list",formula1='"Pendiente,En proceso,Completado,Reprogramado"',allow_blank=True)
ws6.add_data_validation(dv_ec)
for ri,(mes,accion,fecha,norm,dest) in enumerate(cal):
    r=ri+3; bg=GRIS_ALT if ri%2==0 else BLANCO
    for c in range(1,len(HW_CAL)+1):
        cell=ws6.cell(r,c); cell.fill=fill(bg); cell.font=cell_font(); cell.alignment=left(); cell.border=border_thin()
    ws6.cell(r,1).value=mes; ws6.cell(r,2).value=accion
    ws6.cell(r,3).value=fecha; ws6.cell(r,3).number_format="DD/MM/YYYY"
    ws6.cell(r,4).value=norm; ws6.cell(r,5).value=dest; ws6.cell(r,6).value="Pendiente"
    dv_ec.add(f"F{r}")
    ws6.cell(r,7).value=f'=IF(C{r}="","",C{r}-TODAY())'; ws6.cell(r,7).number_format="0"; ws6.cell(r,7).alignment=center()
    ws6.row_dimensions[r].height=20
add_cond_fmt(ws6,"G3:G30")
ws6.freeze_panes="A3"

# ══════════════════════════════════════════════════════════
# HOJA 7: INSTRUCCIONES
# ══════════════════════════════════════════════════════════
ws7=wb.create_sheet("ℹ️ Instrucciones")
ws7.sheet_view.showGridLines=False
ws7.column_dimensions["A"].width=3; ws7.column_dimensions["B"].width=22; ws7.column_dimensions["C"].width=65; ws7.column_dimensions["D"].width=3

items=[
    ("CÓMO USAR ESTA PLANILLA",None,AZUL_OSC,True,13,True),
    (None,None,None,False,8,False),
    ("📚 Alumnos","Registrá hasta 250 alumnos inclusivos. Listas desplegables en: Turno, Tipo de Adecuación, Grado de Apoyo y Estado.",AZUL_MED,True,10,False),
    ("🤝 Familias","Hasta 750 registros de compromisos con familias (~3 por alumno). Días Restantes y colores de alerta automáticos.",AZUL_MED,True,10,False),
    ("📄 Docs MEC","12 documentos preinsertados según normativa. Actualizá Estado y Fecha Límite. Días Restantes calculados solos.",AZUL_MED,True,10,False),
    ("📊 Seguimiento","Hasta 750 registros de progreso por alumno/materia/período. % de Logro con colores automáticos.",AZUL_MED,True,10,False),
    ("📅 Calendario","18 fechas clave del año escolar 2026 según normativa MEC. Marcá el Estado de cada evento.",AZUL_MED,True,10,False),
    (None,None,None,False,8,False),
    ("CÓDIGO DE COLORES",None,AZUL_MED,True,11,True),
    ("🔴 ROJO","Plazo vencido — acción inmediata",ROJO,True,10,False),
    ("🟠 NARANJA","Vence en ≤15 días — preparar urgente",NARANJA,True,10,False),
    ("🟡 AMARILLO","Vence en 30 días — planificar","7F6000",True,10,False),
    ("🟢 VERDE","Sin urgencia inmediata",VERDE,True,10,False),
    (None,None,None,False,8,False),
    ("NORMATIVA",None,AZUL_OSC,True,11,True),
    ("Ley N° 5136/2013","Ley de Educación Inclusiva — Marco general",AZUL_MED,False,10,False),
    ("Res. MEC 29.664/2012","Integración escolar de personas con discapacidad",AZUL_MED,False,10,False),
    ("Decreto N° 1.350/2019","Reglamentación Ley 5136/2013",AZUL_MED,False,10,False),
    ("Res. MEC 29.451/2010","Adecuaciones curriculares",AZUL_MED,False,10,False),
    ("Ley N° 4962/2013","Carta Orgánica SENADIS",AZUL_MED,False,10,False),
]
ws7.merge_cells("A1:D1"); ws7["A1"].fill=fill(AZUL_FONDO)
r=2
for title,desc,color,bold,size,is_section in items:
    if title and desc:
        ws7.cell(r,2).value=title; ws7.cell(r,2).font=Font(name="Arial",bold=True,color=color,size=size); ws7.cell(r,2).alignment=left()
        ws7.cell(r,3).value=desc;  ws7.cell(r,3).font=Font(name="Arial",size=size); ws7.cell(r,3).alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        ws7.row_dimensions[r].height=28
    elif title:
        ws7.merge_cells(f"B{r}:C{r}")
        ws7.cell(r,2).value=title; ws7.cell(r,2).font=Font(name="Arial",bold=bold,color=color,size=size)
        if is_section: ws7.cell(r,2).fill=fill(AZUL_FONDO)
        ws7.cell(r,2).alignment=left(); ws7.row_dimensions[r].height=22
    else:
        ws7.row_dimensions[r].height=8
    r+=1

out="/sessions/sweet-serene-curie/mnt/outputs/Seguimiento_Inclusivo_Cerritos_2026.xlsx"
wb.save(out)
print("OK:", out)
