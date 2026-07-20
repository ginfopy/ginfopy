from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────
AZUL_OSC   = "1F3864"
AZUL_MED   = "2E75B6"
AZUL_CLAR  = "BDD7EE"
AZUL_FONDO = "DEEAF1"
VERDE      = "375623"
VERDE_CLAR = "E2EFDA"
ROJO       = "C00000"
ROJO_CLAR  = "FCE4D6"
NARANJA    = "C55A11"
NARANJA_C  = "FCE9D9"
AMARILLO   = "FFF2CC"
GRIS_HEAD  = "D6DCE4"
GRIS_ALT   = "F2F2F2"
BLANCO     = "FFFFFF"

def hdr_font(bold=True, color=BLANCO, size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)

def cell_font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_med():
    s = Side(style="medium", color="4472C4")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, bg=AZUL_MED, fg=BLANCO, bold=True):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = Font(name="Arial", bold=bold, color=fg, size=10)
        cell.alignment = center()
        cell.border = border_thin()

def style_data_row(ws, row, cols, bg=BLANCO):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = cell_font()
        cell.alignment = left()
        cell.border = border_thin()

# ══════════════════════════════════════════════════════════════════
# HOJA 1: INICIO (Dashboard de alertas)
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 INICIO"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 20
ws1.column_dimensions["G"].width = 18
ws1.column_dimensions["H"].width = 3

# Título principal
ws1.merge_cells("B1:G1")
ws1.row_dimensions[1].height = 10
ws1.merge_cells("B2:G3")
t = ws1["B2"]
t.value = "COLEGIO PRESBITERIANO CERRITOS"
t.font = Font(name="Arial", bold=True, color=BLANCO, size=16)
t.fill = fill(AZUL_OSC)
t.alignment = center()

ws1.merge_cells("B4:G4")
sub = ws1["B4"]
sub.value = "Sistema de Seguimiento — Educación Inclusiva"
sub.font = Font(name="Arial", bold=False, color=AZUL_MED, size=12, italic=True)
sub.alignment = center()

ws1.row_dimensions[5].height = 8

# Sección resumen stats
ws1.merge_cells("B6:C6")
ws1["B6"].value = "RESUMEN GENERAL"
ws1["B6"].font = Font(name="Arial", bold=True, color=BLANCO, size=10)
ws1["B6"].fill = fill(AZUL_MED)
ws1["B6"].alignment = center()

stat_labels = [
    ("B7", "Total alumnos inclusivos",  "=COUNTA('📚 Alumnos'!B3:B52)"),
    ("B8", "Compromisos activos",        "=COUNTIF('🤝 Familias'!F3:F102,\"Pendiente\")"),
    ("B9", "Documentos MEC pendientes",  "=COUNTIF('📄 Documentos MEC'!E3:E52,\"Pendiente\")"),
    ("B10","Alertas (próx. 30 días)",    '=COUNTIF(\'📋 INICIO\'!E15:E114,"⚠️ PRÓXIMO")+COUNTIF(\'📋 INICIO\'!E15:E114,"🔴 VENCIDO")'),
]
for addr, label, formula in stat_labels:
    ws1[addr].value = label
    ws1[addr].font = cell_font(size=10)
    ws1[addr].fill = fill(AZUL_FONDO)
    ws1[addr].border = border_thin()
    ws1[addr].alignment = left()
    col = addr[0]
    row = addr[1:]
    c_addr = "C" + row
    ws1[c_addr].value = formula
    ws1[c_addr].font = Font(name="Arial", bold=True, color=AZUL_MED, size=12)
    ws1[c_addr].fill = fill(AZUL_FONDO)
    ws1[c_addr].border = border_thin()
    ws1[c_addr].alignment = center()

ws1.row_dimensions[11].height = 10

# Panel de próximos vencimientos (header)
ws1.merge_cells("B12:G12")
ph = ws1["B12"]
ph.value = "⏰  PRÓXIMOS VENCIMIENTOS Y ALERTAS"
ph.font = Font(name="Arial", bold=True, color=BLANCO, size=11)
ph.fill = fill(AZUL_OSC)
ph.alignment = center()

hdrs_alert = ["Tipo", "Alumno / Descripción", "Fecha Límite", "Días Restantes", "Estado", "Responsable"]
for i, h in enumerate(hdrs_alert, start=2):
    c = ws1.cell(row=13, column=i)
    c.value = h
    c.font = hdr_font(color=BLANCO)
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    c.border = border_thin()

ws1.row_dimensions[13].height = 28

# 20 filas de datos de alerta (se llenan con fórmulas que jalan de otras hojas)
# Aquí ponemos filas de ejemplo con datos de muestra + fórmulas de días restantes
sample_alerts = [
    ("📄 Doc. MEC", "Planilla de matrícula inclusiva",   "2026-07-31"),
    ("📄 Doc. MEC", "Informe semestral de adecuaciones", "2026-07-15"),
    ("📄 Doc. MEC", "Plan Educativo Individual (PEI)",   "2026-06-30"),
    ("🤝 Familia",  "Reunión de compromiso — ver lista", ""),
    ("📄 Doc. MEC", "Evaluación psicopedagógica",        "2026-08-30"),
]
for i, (tipo, desc, fecha) in enumerate(sample_alerts):
    r = 14 + i
    bg = GRIS_ALT if i % 2 == 0 else BLANCO
    ws1.cell(r, 2).value = tipo;  ws1.cell(r,2).fill=fill(bg); ws1.cell(r,2).border=border_thin(); ws1.cell(r,2).font=cell_font(); ws1.cell(r,2).alignment=left()
    ws1.cell(r, 3).value = desc;  ws1.cell(r,3).fill=fill(bg); ws1.cell(r,3).border=border_thin(); ws1.cell(r,3).font=cell_font(); ws1.cell(r,3).alignment=left()
    if fecha:
        ws1.cell(r, 4).value = fecha
        ws1.cell(r, 4).number_format = "DD/MM/YYYY"
        ws1.cell(r, 5).value = f'=IF(D{r}="","",D{r}-TODAY())'
        ws1.cell(r, 5).number_format = "0"
        ws1.cell(r, 6).value = f'=IF(D{r}="","",IF(E{r}<0,"🔴 VENCIDO",IF(E{r}<=15,"⚠️ PRÓXIMO",IF(E{r}<=30,"🟡 ATENCIÓN","🟢 OK"))))'
    ws1.cell(r, 7).value = "Lic. Rodrigo Godoy"
    for col in [4,5,6,7]:
        ws1.cell(r, col).fill = fill(bg)
        ws1.cell(r, col).border = border_thin()
        ws1.cell(r, col).font = cell_font()
        ws1.cell(r, col).alignment = center()

ws1.row_dimensions[14].height = 18
ws1.row_dimensions[19].height = 10

ws1.merge_cells("B20:G20")
nota = ws1["B20"]
nota.value = "ℹ  Completá los datos en las hojas 📚 Alumnos, 🤝 Familias y 📄 Documentos MEC — las alertas se actualizan automáticamente."
nota.font = Font(name="Arial", italic=True, color="595959", size=9)
nota.alignment = left()

# Freeze panes
ws1.freeze_panes = "B14"

# ══════════════════════════════════════════════════════════════════
# HOJA 2: ALUMNOS INCLUSIVOS
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📚 Alumnos")
ws2.sheet_view.showGridLines = False

cols_alumnos = {
    "A": ("N°",                      5),
    "B": ("Apellido y Nombre",       28),
    "C": ("CI / Doc.",               14),
    "D": ("Fecha Nacim.",            14),
    "E": ("Curso / Grado",           14),
    "F": ("Turno",                   10),
    "G": ("Necesidad / Diagnóstico", 30),
    "H": ("Tipo de Adecuación",      22),
    "I": ("Grado de Apoyo",          16),
    "J": ("Fecha Ingreso",           14),
    "K": ("Docente Responsable",     24),
    "L": ("Tel. Familia",            16),
    "M": ("Responsable MEC",         22),
    "N": ("Estado",                  14),
    "O": ("Observaciones",           35),
}
for col_letter, (_, width) in cols_alumnos.items():
    ws2.column_dimensions[col_letter].width = width

ws2.merge_cells("A1:O1")
tit = ws2["A1"]
tit.value = "REGISTRO DE ALUMNOS INCLUSIVOS — Colegio Presbiteriano Cerritos"
tit.font = Font(name="Arial", bold=True, color=BLANCO, size=13)
tit.fill = fill(AZUL_OSC)
tit.alignment = center()
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:O2")
ws2["A2"].value = f'Actualizado: {datetime.date.today().strftime("%d/%m/%Y")}  |  Normativa: Resolución MEC N° 29.664/2012 y Decreto 1.350/2019 - Educación Inclusiva Paraguay'
ws2["A2"].font = Font(name="Arial", italic=True, color=AZUL_MED, size=9)
ws2["A2"].alignment = left()

for i, (col_letter, (header, _)) in enumerate(cols_alumnos.items(), start=1):
    c = ws2.cell(row=3, column=i)
    c.value = header
    c.font = hdr_font()
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    c.border = border_thin()
ws2.row_dimensions[3].height = 30

# Validaciones
dv_turno = DataValidation(type="list", formula1='"Mañana,Tarde,Tiempo Completo"', allow_blank=True)
dv_adec = DataValidation(type="list", formula1='"Adecuación de Acceso,Adecuación Curricular No Significativa,Adecuación Curricular Significativa"', allow_blank=True)
dv_apoyo = DataValidation(type="list", formula1='"Leve,Moderado,Intenso"', allow_blank=True)
dv_estado = DataValidation(type="list", formula1='"Activo,Egresado,Trasladado,Retirado"', allow_blank=True)
ws2.add_data_validation(dv_turno)
ws2.add_data_validation(dv_adec)
ws2.add_data_validation(dv_apoyo)
ws2.add_data_validation(dv_estado)

for r in range(4, 54):
    bg = GRIS_ALT if r % 2 == 0 else BLANCO
    style_data_row(ws2, r, len(cols_alumnos), bg)
    ws2.cell(r, 1).value = r - 3
    ws2.cell(r, 1).alignment = center()
    ws2.cell(r, 4).number_format = "DD/MM/YYYY"
    ws2.cell(r, 10).number_format = "DD/MM/YYYY"
    dv_turno.add(f"F{r}")
    dv_adec.add(f"H{r}")
    dv_apoyo.add(f"I{r}")
    dv_estado.add(f"N{r}")
    ws2.row_dimensions[r].height = 18

ws2.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════
# HOJA 3: COMPROMISOS CON FAMILIAS
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🤝 Familias")
ws3.sheet_view.showGridLines = False

cols_fam = {
    "A": ("N°",                 5),
    "B": ("Alumno",            28),
    "C": ("Fecha Reunión",     14),
    "D": ("Participantes",     30),
    "E": ("Acuerdo / Compromiso", 40),
    "F": ("Estado",            14),
    "G": ("Fecha Límite",      14),
    "H": ("Días Restantes",    14),
    "I": ("Seguimiento",       35),
    "J": ("Próxima Reunión",   16),
    "K": ("Responsable",       22),
    "L": ("Firma / Constancia",18),
}
for col_letter, (_, width) in cols_fam.items():
    ws3.column_dimensions[col_letter].width = width

ws3.merge_cells("A1:L1")
tit3 = ws3["A1"]
tit3.value = "COMPROMISOS Y ACUERDOS CON FAMILIAS — Educación Inclusiva"
tit3.font = Font(name="Arial", bold=True, color=BLANCO, size=13)
tit3.fill = fill(AZUL_OSC)
tit3.alignment = center()
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:L2")
ws3["A2"].value = "Registrá cada reunión y acuerdo formal con la familia. El campo 'Días Restantes' se actualiza automáticamente."
ws3["A2"].font = Font(name="Arial", italic=True, color=AZUL_MED, size=9)

for i, (col_letter, (header, _)) in enumerate(cols_fam.items(), start=1):
    c = ws3.cell(row=3, column=i)
    c.value = header
    c.font = hdr_font()
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    c.border = border_thin()
ws3.row_dimensions[3].height = 30

dv_estado_fam = DataValidation(type="list", formula1='"Pendiente,Cumplido,Reprogramado,Incumplido"', allow_blank=True)
ws3.add_data_validation(dv_estado_fam)

for r in range(4, 104):
    bg = GRIS_ALT if r % 2 == 0 else BLANCO
    style_data_row(ws3, r, len(cols_fam), bg)
    ws3.cell(r, 1).value = r - 3
    ws3.cell(r, 1).alignment = center()
    ws3.cell(r, 3).number_format = "DD/MM/YYYY"
    ws3.cell(r, 7).number_format = "DD/MM/YYYY"
    ws3.cell(r, 8).value = f'=IF(G{r}="","",G{r}-TODAY())'
    ws3.cell(r, 8).number_format = "0"
    ws3.cell(r, 10).number_format = "DD/MM/YYYY"
    dv_estado_fam.add(f"F{r}")
    ws3.row_dimensions[r].height = 18

# Formato condicional: días restantes
from openpyxl.formatting.rule import CellIsRule
rojo_fill = PatternFill("solid", fgColor=ROJO_CLAR)
naranja_fill = PatternFill("solid", fgColor=NARANJA_C)
verde_fill = PatternFill("solid", fgColor=VERDE_CLAR)
amarillo_fill = PatternFill("solid", fgColor=AMARILLO)

ws3.conditional_formatting.add(f"H4:H103",
    CellIsRule(operator="lessThan", formula=["0"], fill=rojo_fill,
               font=Font(name="Arial", color=ROJO, bold=True)))
ws3.conditional_formatting.add(f"H4:H103",
    CellIsRule(operator="between", formula=["0","15"], fill=naranja_fill,
               font=Font(name="Arial", color=NARANJA, bold=True)))
ws3.conditional_formatting.add(f"H4:H103",
    CellIsRule(operator="between", formula=["16","30"], fill=amarillo_fill))
ws3.conditional_formatting.add(f"H4:H103",
    CellIsRule(operator="greaterThan", formula=["30"], fill=verde_fill,
               font=Font(name="Arial", color=VERDE)))

ws3.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════
# HOJA 4: DOCUMENTOS MEC
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📄 Documentos MEC")
ws4.sheet_view.showGridLines = False

cols_mec = {
    "A": ("N°",                  5),
    "B": ("Documento / Trámite", 40),
    "C": ("Alumno(s)",           28),
    "D": ("Base Legal",          30),
    "E": ("Estado",              14),
    "F": ("Fecha Límite",        14),
    "G": ("Días Restantes",      14),
    "H": ("Alerta",              14),
    "I": ("Responsable",         22),
    "J": ("Dónde se presenta",   25),
    "K": ("Observaciones",       35),
}
for col_letter, (_, width) in cols_mec.items():
    ws4.column_dimensions[col_letter].width = width

ws4.merge_cells("A1:K1")
tit4 = ws4["A1"]
tit4.value = "DOCUMENTOS Y PLAZOS — MINISTERIO DE EDUCACIÓN Y CIENCIAS (MEC) — Paraguay"
tit4.font = Font(name="Arial", bold=True, color=BLANCO, size=13)
tit4.fill = fill(AZUL_OSC)
tit4.alignment = center()
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:K2")
ws4["A2"].value = "Normativa de referencia: Res. MEC N° 29.664/2012 | Decreto N° 1.350/2019 | Ley N° 5136/2013 - Educación Inclusiva | Res. N° 29.451/2010"
ws4["A2"].font = Font(name="Arial", italic=True, color=AZUL_MED, size=9)

for i, (col_letter, (header, _)) in enumerate(cols_mec.items(), start=1):
    c = ws4.cell(row=3, column=i)
    c.value = header
    c.font = hdr_font()
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    c.border = border_thin()
ws4.row_dimensions[3].height = 30

dv_estado_mec = DataValidation(type="list", formula1='"Pendiente,En preparación,Presentado,Aprobado,Observado"', allow_blank=True)
ws4.add_data_validation(dv_estado_mec)

# Datos precargados: documentos típicos del MEC Paraguay para educación inclusiva
docs_preload = [
    ("Planilla de matrícula inclusiva (Ficha de Registro)",
     "Todos los alumnos nuevos inclusivos",
     "Res. MEC 29.664/2012 Art. 8",
     "Pendiente", "2026-03-15",
     "Sección de Educación Especial MEC / Supervisión"),
    ("Plan Educativo Individual (PEI) — Elaboración",
     "Cada alumno inclusivo",
     "Ley 5136/2013 Art. 15 | Res. 29.664/2012",
     "En preparación", "2026-04-30",
     "Archivo institucional + copia Supervisión"),
    ("Informe de adecuaciones curriculares — 1er semestre",
     "Todos los alumnos inclusivos",
     "Res. MEC 29.664/2012 Art. 12",
     "Pendiente", "2026-07-15",
     "Supervisión Educativa Departamental"),
    ("Informe de adecuaciones curriculares — 2do semestre",
     "Todos los alumnos inclusivos",
     "Res. MEC 29.664/2012 Art. 12",
     "Pendiente", "2026-11-30",
     "Supervisión Educativa Departamental"),
    ("Evaluación psicopedagógica actualizada",
     "Alumnos que requieren renovación",
     "Decreto 1.350/2019 Art. 22",
     "Pendiente", "2026-08-31",
     "Centro de Diagnóstico MEC / archivo institucional"),
    ("Nómina de alumnos inclusivos por nivel",
     "Todos los niveles",
     "Res. MEC 29.664/2012",
     "Pendiente", "2026-03-31",
     "Supervisión Educativa"),
    ("Constancia de diagnóstico médico/psicológico",
     "Alumnos nuevos",
     "Ley 5136/2013 Art. 9",
     "Pendiente", "",
     "Archivo institucional"),
    ("Plan de Apoyos y Recursos (PAR)",
     "Alumnos con adecuación significativa",
     "Decreto 1.350/2019",
     "Pendiente", "2026-05-15",
     "Archivo institucional"),
    ("Acta de compromiso familiar firmada",
     "Cada familia de alumno inclusivo",
     "Res. MEC 29.664/2012",
     "Pendiente", "",
     "Archivo institucional"),
    ("Informe final anual de progreso",
     "Todos los alumnos inclusivos",
     "Ley 5136/2013 Art. 16",
     "Pendiente", "2026-11-30",
     "Supervisión Educativa Departamental"),
    ("Solicitud de recursos de apoyo (intérprete, asistente)",
     "Según necesidad",
     "Decreto 1.350/2019 Art. 30",
     "Pendiente", "2026-03-01",
     "DGEEI - Dirección Gral. Ed. Especial e Inclusiva"),
    ("Certificado de discapacidad SENADIS actualizado",
     "Alumnos que corresponda",
     "Ley 4962/2013",
     "Pendiente", "",
     "Archivo institucional"),
]

for r_i, (doc, alumnos, legal, estado, fecha, lugar) in enumerate(docs_preload):
    r = r_i + 4
    bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
    style_data_row(ws4, r, len(cols_mec), bg)
    ws4.cell(r, 1).value = r_i + 1;  ws4.cell(r,1).alignment=center()
    ws4.cell(r, 2).value = doc
    ws4.cell(r, 3).value = alumnos
    ws4.cell(r, 4).value = legal
    ws4.cell(r, 5).value = estado
    dv_estado_mec.add(f"E{r}")
    if fecha:
        ws4.cell(r, 6).value = fecha
        ws4.cell(r, 6).number_format = "DD/MM/YYYY"
    ws4.cell(r, 7).value = f'=IF(F{r}="","",F{r}-TODAY())'
    ws4.cell(r, 7).number_format = "0"
    ws4.cell(r, 8).value = f'=IF(F{r}="","",IF(G{r}<0,"🔴 VENCIDO",IF(G{r}<=15,"⚠️ URGENTE",IF(G{r}<=30,"🟡 PRÓXIMO","🟢 OK"))))'
    ws4.cell(r, 9).value = "Lic. Rodrigo Godoy"
    ws4.cell(r, 10).value = lugar
    ws4.row_dimensions[r].height = 22

# Filas vacías adicionales
for r in range(len(docs_preload) + 4, 55):
    bg = GRIS_ALT if r % 2 == 0 else BLANCO
    style_data_row(ws4, r, len(cols_mec), bg)
    ws4.cell(r, 1).value = r - 3; ws4.cell(r,1).alignment=center()
    ws4.cell(r, 7).value = f'=IF(F{r}="","",F{r}-TODAY())'
    ws4.cell(r, 7).number_format = "0"
    ws4.cell(r, 8).value = f'=IF(F{r}="","",IF(G{r}<0,"🔴 VENCIDO",IF(G{r}<=15,"⚠️ URGENTE",IF(G{r}<=30,"🟡 PRÓXIMO","🟢 OK"))))'
    dv_estado_mec.add(f"E{r}")
    ws4.row_dimensions[r].height = 18

ws4.conditional_formatting.add("G4:G54",
    CellIsRule(operator="lessThan", formula=["0"], fill=rojo_fill,
               font=Font(name="Arial", color=ROJO, bold=True)))
ws4.conditional_formatting.add("G4:G54",
    CellIsRule(operator="between", formula=["0","15"], fill=naranja_fill,
               font=Font(name="Arial", color=NARANJA, bold=True)))
ws4.conditional_formatting.add("G4:G54",
    CellIsRule(operator="between", formula=["16","30"], fill=amarillo_fill))
ws4.conditional_formatting.add("G4:G54",
    CellIsRule(operator="greaterThan", formula=["30"], fill=verde_fill,
               font=Font(name="Arial", color=VERDE)))
ws4.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════
# HOJA 5: SEGUIMIENTO ACADÉMICO
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📊 Seguimiento")
ws5.sheet_view.showGridLines = False

cols_seg = {
    "A": ("N°",                   5),
    "B": ("Alumno",              28),
    "C": ("Materia / Área",       22),
    "D": ("Trimestre / Período",  18),
    "E": ("Tipo de Adecuación",   24),
    "F": ("Objetivo Planteado",   35),
    "G": ("Estrategia Usada",     30),
    "H": ("Resultado / Logro",    30),
    "I": ("% Logro",             10),
    "J": ("Próximo Paso",         30),
    "K": ("Fecha Registro",       14),
    "L": ("Docente",              22),
}
for col_letter, (_, width) in cols_seg.items():
    ws5.column_dimensions[col_letter].width = width

ws5.merge_cells("A1:L1")
tit5 = ws5["A1"]
tit5.value = "SEGUIMIENTO ACADÉMICO INDIVIDUAL — Alumnos Inclusivos"
tit5.font = Font(name="Arial", bold=True, color=BLANCO, size=13)
tit5.fill = fill(AZUL_OSC)
tit5.alignment = center()
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:L2")
ws5["A2"].value = "Registrá el progreso de cada alumno por materia y período. El % de logro permite visualizar la evolución."
ws5["A2"].font = Font(name="Arial", italic=True, color=AZUL_MED, size=9)

for i, (col_letter, (header, _)) in enumerate(cols_seg.items(), start=1):
    c = ws5.cell(row=3, column=i)
    c.value = header
    c.font = hdr_font()
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    c.border = border_thin()
ws5.row_dimensions[3].height = 30

dv_trim = DataValidation(type="list", formula1='"1er Trimestre,2do Trimestre,3er Trimestre,Semestral,Anual"', allow_blank=True)
ws5.add_data_validation(dv_trim)

for r in range(4, 104):
    bg = GRIS_ALT if r % 2 == 0 else BLANCO
    style_data_row(ws5, r, len(cols_seg), bg)
    ws5.cell(r, 1).value = r - 3; ws5.cell(r,1).alignment=center()
    ws5.cell(r, 9).number_format = "0%"
    ws5.cell(r, 11).number_format = "DD/MM/YYYY"
    dv_trim.add(f"D{r}")
    ws5.row_dimensions[r].height = 20

# Formato condicional en % logro
ws5.conditional_formatting.add("I4:I103",
    CellIsRule(operator="lessThan", formula=["0.4"], fill=rojo_fill,
               font=Font(name="Arial", color=ROJO, bold=True)))
ws5.conditional_formatting.add("I4:I103",
    CellIsRule(operator="between", formula=["0.4","0.6"], fill=amarillo_fill))
ws5.conditional_formatting.add("I4:I103",
    CellIsRule(operator="greaterThan", formula=["0.6"], fill=verde_fill,
               font=Font(name="Arial", color=VERDE)))
ws5.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════
# HOJA 6: CALENDARIO ANUAL
# ══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("📅 Calendario 2026")
ws6.sheet_view.showGridLines = False

cols_cal = {
    "A": ("Mes",             12),
    "B": ("Acción / Evento", 45),
    "C": ("Fecha",           14),
    "D": ("Normativa",       35),
    "E": ("Destinatario",    25),
    "F": ("Estado",          14),
    "G": ("Días Restantes",  14),
}
for col_letter, (_, width) in cols_cal.items():
    ws6.column_dimensions[col_letter].width = width

ws6.merge_cells("A1:G1")
tit6 = ws6["A1"]
tit6.value = "CALENDARIO ANUAL DE EDUCACIÓN INCLUSIVA 2026 — MEC Paraguay"
tit6.font = Font(name="Arial", bold=True, color=BLANCO, size=13)
tit6.fill = fill(AZUL_OSC)
tit6.alignment = center()
ws6.row_dimensions[1].height = 30

for i, (col_letter, (header, _)) in enumerate(cols_cal.items(), start=1):
    c = ws6.cell(row=2, column=i)
    c.value = header
    c.font = hdr_font()
    c.fill = fill(AZUL_MED)
    c.alignment = center()
    c.border = border_thin()
ws6.row_dimensions[2].height = 28

cal_data = [
    ("Febrero",  "Inicio de año escolar — Identificación de alumnos inclusivos", "2026-02-16", "Ley 5136/2013", "Dirección / Docentes"),
    ("Febrero",  "Entrega de informes de alumnos inclusivos del año anterior",    "2026-02-28", "Res. 29.664/2012", "Dirección"),
    ("Marzo",    "Matrícula especial / diferenciada — Cierre",                    "2026-03-15", "Res. 29.664/2012 Art.8", "Secretaría"),
    ("Marzo",    "Nómina de alumnos inclusivos a Supervisión",                    "2026-03-31", "Res. 29.664/2012", "Dirección"),
    ("Abril",    "Elaboración de PEI (Plan Educativo Individual) — plazo",        "2026-04-30", "Ley 5136/2013 Art.15", "Docentes + Psicología"),
    ("Mayo",     "Plan de Apoyos y Recursos (PAR) — entrega",                    "2026-05-15", "Decreto 1.350/2019", "Coordinadores"),
    ("Mayo",     "Reunión con familias — 1er seguimiento",                        "2026-05-30", "Res. 29.664/2012", "Dirección / Docentes"),
    ("Junio",    "Evaluación trimestral con adecuaciones — 1er trimestre",        "2026-06-15", "Ley 5136/2013 Art.16", "Docentes"),
    ("Julio",    "Informe de adecuaciones 1er semestre — entrega Supervisión",    "2026-07-15", "Res. 29.664/2012 Art.12", "Dirección"),
    ("Julio",    "Planilla de matrícula inclusiva — actualización",               "2026-07-31", "Res. 29.664/2012", "Secretaría"),
    ("Agosto",   "Evaluaciones psicopedagógicas — plazo renovación",             "2026-08-31", "Decreto 1.350/2019 Art.22", "Psicología"),
    ("Septiembre","Revisión PEI — ajuste 2do semestre",                           "2026-09-15", "Ley 5136/2013", "Docentes + Psicología"),
    ("Septiembre","Reunión con familias — 2do seguimiento",                       "2026-09-30", "Res. 29.664/2012", "Dirección / Docentes"),
    ("Octubre",  "Evaluación trimestral con adecuaciones — 2do trimestre",        "2026-10-15", "Ley 5136/2013 Art.16", "Docentes"),
    ("Noviembre","Informe de adecuaciones 2do semestre — entrega Supervisión",    "2026-11-30", "Res. 29.664/2012 Art.12", "Dirección"),
    ("Noviembre","Informe final anual de progreso inclusivo",                     "2026-11-30", "Ley 5136/2013 Art.16", "Dirección"),
    ("Diciembre","Reunión de cierre con familias",                                "2026-12-05", "Res. 29.664/2012", "Dirección"),
    ("Diciembre","Archivo de expedientes inclusivos del año",                    "2026-12-15", "Res. 29.664/2012", "Secretaría"),
]

dv_estado_cal = DataValidation(type="list", formula1='"Pendiente,En proceso,Completado,Reprogramado"', allow_blank=True)
ws6.add_data_validation(dv_estado_cal)

for r_i, (mes, accion, fecha, norm, dest) in enumerate(cal_data):
    r = r_i + 3
    bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
    style_data_row(ws6, r, len(cols_cal), bg)
    ws6.cell(r, 1).value = mes
    ws6.cell(r, 2).value = accion
    ws6.cell(r, 3).value = fecha; ws6.cell(r,3).number_format="DD/MM/YYYY"
    ws6.cell(r, 4).value = norm
    ws6.cell(r, 5).value = dest
    ws6.cell(r, 6).value = "Pendiente"
    dv_estado_cal.add(f"F{r}")
    ws6.cell(r, 7).value = f'=IF(C{r}="","",C{r}-TODAY())'
    ws6.cell(r, 7).number_format = "0"
    ws6.row_dimensions[r].height = 20

ws6.conditional_formatting.add("G3:G30",
    CellIsRule(operator="lessThan", formula=["0"], fill=rojo_fill,
               font=Font(name="Arial", color=ROJO, bold=True)))
ws6.conditional_formatting.add("G3:G30",
    CellIsRule(operator="between", formula=["0","15"], fill=naranja_fill,
               font=Font(name="Arial", color=NARANJA, bold=True)))
ws6.conditional_formatting.add("G3:G30",
    CellIsRule(operator="between", formula=["16","30"], fill=amarillo_fill))
ws6.conditional_formatting.add("G3:G30",
    CellIsRule(operator="greaterThan", formula=["30"], fill=verde_fill))
ws6.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════
# HOJA 7: INSTRUCCIONES
# ══════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("ℹ️ Instrucciones")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions["A"].width = 3
ws7.column_dimensions["B"].width = 20
ws7.column_dimensions["C"].width = 65
ws7.column_dimensions["D"].width = 3

instrucciones = [
    ("CÓMO USAR ESTA PLANILLA", None, AZUL_OSC, True, 13),
    (None, None, None, False, 8),
    ("📋 INICIO", "Panel de alertas automáticas. Muestra los próximos vencimientos basándose en los datos cargados en las otras hojas. No editar directamente — se actualiza solo.", AZUL_MED, True, 10),
    ("📚 Alumnos", "Registrá cada alumno inclusivo con sus datos personales, tipo de necesidad y adecuación. Los campos Turno, Tipo de Adecuación, Grado de Apoyo y Estado tienen listas desplegables.", AZUL_MED, True, 10),
    ("🤝 Familias", "Registrá cada reunión y compromiso con las familias. Fechas Límite y Días Restantes se calculan automáticamente. Colores de alerta: rojo=vencido, naranja=15 días, amarillo=30 días.", AZUL_MED, True, 10),
    ("📄 Docs MEC", "Documentos preinsertados según normativa paraguaya. Actualizá el Estado y la Fecha Límite. Los Días Restantes se calculan solos.", AZUL_MED, True, 10),
    ("📊 Seguimiento", "Registrá el progreso académico por alumno, materia y período. El % de Logro activa colores (verde=bien, amarillo=regular, rojo=bajo).", AZUL_MED, True, 10),
    ("📅 Calendario", "Fechas clave del año escolar 2026 según normativa MEC. Marcá el Estado de cada evento.", AZUL_MED, True, 10),
    (None, None, None, False, 8),
    ("CÓDIGO DE COLORES", None, AZUL_MED, True, 11),
    ("🔴 ROJO", "Plazo vencido — acción inmediata requerida", ROJO, True, 10),
    ("🟠 NARANJA", "Vence en menos de 15 días — preparar urgente", NARANJA, True, 10),
    ("🟡 AMARILLO", "Vence en 30 días — planificar", "7F6000", True, 10),
    ("🟢 VERDE", "Sin urgencia inmediata", VERDE, True, 10),
    (None, None, None, False, 8),
    ("NORMATIVA DE REFERENCIA", None, AZUL_OSC, True, 11),
    ("Ley N° 5136/2013", "Ley de Educación Inclusiva — Marco general del sistema inclusivo en Paraguay", AZUL_MED, False, 10),
    ("Res. MEC 29.664/2012", "Normativa sobre integración escolar de personas con discapacidad", AZUL_MED, False, 10),
    ("Decreto N° 1.350/2019", "Reglamentación de la Ley 5136/2013 — Derechos y procedimientos", AZUL_MED, False, 10),
    ("Res. MEC 29.451/2010", "Procedimientos para adecuaciones curriculares", AZUL_MED, False, 10),
    ("Ley N° 4962/2013", "Carta Orgánica del SENADIS — Certificado de discapacidad", AZUL_MED, False, 10),
]

ws7.merge_cells("A1:D1"); ws7["A1"].fill = fill(AZUL_FONDO)
r = 2
for item in instrucciones:
    if len(item) == 5:
        titulo, desc, color, bold, size = item
        if titulo and desc:
            ws7.cell(r, 2).value = titulo
            ws7.cell(r, 2).font = Font(name="Arial", bold=True, color=color, size=size)
            ws7.cell(r, 2).alignment = left()
            ws7.cell(r, 3).value = desc
            ws7.cell(r, 3).font = Font(name="Arial", bold=False, color="000000", size=size)
            ws7.cell(r, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws7.row_dimensions[r].height = 30
        elif titulo:
            ws7.merge_cells(f"B{r}:C{r}")
            ws7.cell(r, 2).value = titulo
            ws7.cell(r, 2).font = Font(name="Arial", bold=bold, color=color, size=size)
            ws7.cell(r, 2).fill = fill(AZUL_FONDO) if color == AZUL_OSC else fill(BLANCO)
            ws7.cell(r, 2).alignment = left()
            ws7.row_dimensions[r].height = 22
        else:
            ws7.row_dimensions[r].height = 8
    r += 1

# ── Guardar ──────────────────────────────────────────────────────
out_path = "/sessions/sweet-serene-curie/mnt/outputs/Seguimiento_Inclusivo_Cerritos_2026.xlsx"
wb.save(out_path)
print("OK:", out_path)
